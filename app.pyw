from openpyxl import load_workbook
import customtkinter as ctk
from PIL import Image, ImageDraw, ImageFont
import textwrap
from datetime import datetime
from os import startfile

wb = load_workbook("\\\\cim-nas\\Servidor_Recursos_Humanos\\RECURSOS HUMANOS 2024\\FORMATOS RH\\PLANTILLA 2024.xlsx")
ws = wb["activos"]


#  ---Principal function that executes on click
def on_button_click():
    searchNumber(False)
def on_button_click2():
    searchNumber(True)

def searchNumber(paper):
    try:
        value = int(text_box.get())
    except ValueError:
        error.configure(text="Ingrese el No. de empleado")
        return
    column = "D"
    row = None
    for cell in ws[column]:
        if cell.value == value:
            row = cell.row
            searchData(row, paper)
    if not row:
        error.configure(text="No se encontro empleado")


def searchData(row, paper):

    number = str(ws[f"D{row}"].value if ws[f"D{row}"].value != None else "")
    name = str(ws[f"E{row}"].value if ws[f"E{row}"].value != None else "")
    NSS = str(ws[f"I{row}"].value if ws[f"I{row}"].value != None else "")
    CURP = str(ws[f"K{row}"].value if ws[f"K{row}"].value != None else "")
    RFC = str(ws[f"L{row}"].value if ws[f"L{row}"].value != None else "")
    blood = str(ws[f"N{row}"].value if ws[f"N{row}"].value != None else "")
    account = str(ws[f"T{row}"].value if ws[f"T{row}"].value != None else "")
    emergencyName = str(ws[f"X{row}"].value if ws[f"X{row}"].value != None else "")
    emergencyNumber = str(ws[f"Y{row}"].value if ws[f"Y{row}"].value != None else "")
    date = str(ws[f"AE{row}"].value if ws[f"AE{row}"].value != None else "")
    job = str(ws[f"AB{row}"].value if ws[f"AB{row}"].value != None else "")

    date = datetime.strptime(date, "%Y-%m-%d %H:%M:%S")
    date = date.strftime("%d/%m/%Y")

    generateImage(
        number,
        name,
        NSS,
        CURP,
        RFC,
        account,
        emergencyName,
        emergencyNumber,
        blood,
        job,
        date,
        paper
    )


def generateImage(
    number,
    name,
    NSS,
    CURP,
    RFC,
    account,
    emergencyName,
    emergencyNumber,
    blood,
    job,
    date,
    paper
):
    error.configure(text="")
    # resources
    font1 = ImageFont.truetype("fonts\\OpenSans-Bold.ttf", 25)
    font2 = ImageFont.truetype("fonts\\OpenSans-Bold.ttf", 35)
    font3 = ImageFont.truetype("fonts\\OpenSans-Regular.ttf", 30)
    font4 = ImageFont.truetype("fonts\\OpenSans-Bold.ttf", 20)
    font5 = ImageFont.truetype("fonts\\OpenSans-Bold.ttf", 25)
    font6 = ImageFont.truetype("fonts\\OpenSans-Bold.ttf", 33)
    font7 = ImageFont.truetype("fonts\\OpenSans-Regular.ttf", 33)

    black = (0, 0, 0)
    white = (255, 255, 255)
    # ---first baseImage
    baseImage = Image.open("img\\frontbase.png").convert("RGBA")
    draw = ImageDraw.Draw(baseImage)
    try:
        photo = Image.open(f"\\\\cim-nas\\Servidor_Recursos_Humanos\\RECURSOS HUMANOS 2024\\CREDENCIALES\\FOTOS PARA CREDENCIALES\\Foto Sin fondo\\{number}.png")
    except FileNotFoundError:
        return error.configure(text="No se encontro la foto")

    original_width, original_height = photo.size
    new_height = 330
    new_width = int(new_height * original_width / original_height)
    photo = photo.resize((new_width, new_height))
    centerx = int((baseImage.width - new_width) / 2)
    baseImage.paste(photo, (centerx, 280), photo)

    frontImage = Image.open("img\\front.png")
    baseImage.paste(frontImage, (0, 0), frontImage)

    lines = textwrap.wrap(name, width=25)
    if len(lines) == 1:
        y_text = 640
    else:
        y_text = 610

    for line in lines:
        centerx = (baseImage.width - (draw.textlength(line, font=font2))) / 2
        draw.text((centerx, y_text), line, font=font2, fill=black)
        y_text = 645

    centerx = (baseImage.width - (draw.textlength(job, font=font3))) / 2
    draw.text((centerx, 695), job, font=font3, fill=black)

    centerx = (baseImage.width - (draw.textlength(number, font=font1))) / 2
    draw.text((centerx, 791), number, font=font1, fill=white)

    centerx = (baseImage.width - (draw.textlength(date, font=font5))) / 2
    draw.text((centerx, 889), date, font=font5, fill=white)
    baseImage = baseImage.convert("RGB")
    baseImage.save("result\\front.jpg") 

    # -----------------second baseImage --------------------------------
    baseImage = Image.open("img\\back.png")
    draw = ImageDraw.Draw(baseImage)

    draw.text((70, 150), "CTA:", font=font6, fill=black)
    draw.text((70, 240), "NSS:", font=font6, fill=black)
    draw.text((70, 330), "T.SANGRE:", font=font6, fill=black)
    draw.text((70, 420), "RFC:", font=font6, fill=black)
    draw.text((70, 510), "CURP:", font=font6, fill=black)

    draw.text((150, 150), account, font=font7, fill=black)
    draw.text((150, 240), NSS, font=font7, fill=black)
    draw.text((243, 330), blood, font=font7, fill=black)
    draw.text((148, 420), RFC, font=font7, fill=black)
    draw.text((173, 510), CURP, font=font7, fill=black)

    centerx = (baseImage.width - (draw.textlength("En caso de emergencia llamar a:", font=font1))) / 2
    draw.text((centerx, 600), "En caso de emergencia llamar a:", font=font1, fill=black)

    centerx = (baseImage.width - (draw.textlength(f"{emergencyName} / {emergencyNumber}", font=font7))) / 2
    draw.text((centerx, 630), f"{emergencyName} / {emergencyNumber}", font=font7, fill=black)

    baseImage = baseImage.convert("RGB")
    baseImage.save("result\\back.jpg") 

    if paper is True:
        baseImage = Image.open("img\\paper.jpg")
        firstImage = Image.open("result\\front.jpg")
        secondImage = Image.open("result\\back.jpg")

        baseImage.paste(firstImage, (0, 0))
        baseImage.paste(secondImage, (610, 0))

        baseImage = baseImage
        baseImage.save(f"result\\paper\\{number}.jpg") 
        startfile(f"result\\paper\\{number}.jpg") 
    else:
        startfile("result\\front.jpg") 
        startfile("result\\back.jpg")
    


# ---Make the windows and widgets
app = ctk.CTk()
app.geometry("300x200")
app.title("Credenciales")
app.resizable(False, False)
ctk.set_appearance_mode("light")

frame = ctk.CTkFrame(app, bg_color="#FEFAFA", fg_color="#FEFAFA")
frame.pack(fill="both", expand=True)

label = ctk.CTkLabel(master=frame, text="Numero de empleado:")
label.pack(pady=10)

text_box = ctk.CTkEntry(master=frame)
text_box.pack(pady=10)

button = ctk.CTkButton(master=frame, text="Imprimir tarjeta", command=on_button_click)
button2 = ctk.CTkButton(master=frame, text="Imprimir en hoja", command=on_button_click2)

button.configure(fg_color="#B42725", hover_color="#871e1c")
button.pack(pady=10)
button2.configure(fg_color="#B42725", hover_color="#871e1c")
button2.pack(pady=10)
error = ctk.CTkLabel(master=frame, text="")
error.pack(pady=10)

app.mainloop()
